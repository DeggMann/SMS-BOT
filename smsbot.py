import openpyxl
import pandas as pd
import filehandler as fh
# test code for adding a new route to the route_sheet.xlsx file


def start():
    print("Enter the order details to add a new route.")

    order_number = input("Enter order number: ")
    origin = input("Enter origin: ")
    destination = input("Enter destination: ")

    route_data = [order_number, origin, origin, destination, "In Transit"]
    fh.save_data(route_data)
    print("Successfully added the new route data!")

def update():
    file_name = "route_sheet.xlsx"

    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    
    order_number = input("Enter order number to update: ")
    current_location = input("Enter current location: ")
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == order_number:
            row[2].value = current_location

            if current_location == row[3].value:
                row[4].value = "Delivered"

            workbook.save(file_name)
            print(f"Successfully updated order {order_number} with current location: {current_location}")
            return
    print(f"Order number {order_number} not found.")

def view():
    print("text 1 to view all orders or 2 to view a specific order")
    choice = input("choose option: ")
    file_name = "route_sheet.xlsx"

    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    # 1. Load the Excel file FIRST, before the if/elif blocks, 
    # so 'df' is available to both options.
    df = pd.read_excel(file_name) 
    
    if choice == "1":
        print(df.to_markdown(index=False))
        
    elif choice == "2":
        order_number = input("Enter order number to view: ")
        
        # 2. Make sure 'target_column' matches your actual Excel column name (e.g., 'ID' or 'Order Number')
        target_column = 'Route number' 
        
        # Now 'df' will successfully filter without crashing
        specific_order = df[df[target_column].astype(str) == str(order_number)]
        
        if not specific_order.empty:
            print(specific_order.to_markdown(index=False))
        else:
            print(f"No order found with ID: {order_number}")
            
    else:
        print("Invalid choice. Please enter 1 or 2.")

